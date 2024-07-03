from openpyxl import Workbook,load_workbook
from openpyxl.styles import Alignment, Font, Border, Side,PatternFill
import os
import re
import pandas as pd
import xlsxwriter
import openpyxl


def create_excel_template(folder, excel_folder):
    # Create a new workbook and select the active worksheet
    new_wb = Workbook()
    ws = new_wb.active
    ws.title = "Financial Information"

    # Set column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 50

    # Set bold font for headers
    bold_font = Font(bold=True)
    
    # Set alignment for headers
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')

    # Set borders
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    thick_border = Border(left=Side(style='thick'), 
                          right=Side(style='thick'), 
                          top=Side(style='thick'), 
                          bottom=Side(style='thick'))

    # Fill for headers
    light_grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Add headers and format them
    ws["A1"].value = "Full Name of the company"
    ws["A1"].font = bold_font
    ws["A1"].alignment = center_alignment
    ws["A1"].border = thin_border

    # Financial Information Section
    ws["A5"].value = "A. Company wide Financial Information"
    ws["A5"].font = bold_font
    ws["A5"].alignment = center_alignment
    ws["A5"].border = thin_border

    ws["A6"].value = "Particulars"
    ws["A6"].font = bold_font
    ws["A6"].alignment = center_alignment
    ws["A6"].border = thick_border
    ws["A6"].fill = light_grey_fill
    ws["B6"].border = thick_border
    ws["B6"].fill = light_grey_fill
    ws["C6"].border = thick_border
    ws["C6"].fill = light_grey_fill
    ws["D6"].border = thin_border
    ws["E6"].border = thin_border

    ws["B7"].value = "INR(values in Crores/Lakhs)"
    ws["B7"].font = bold_font
    ws["B7"].alignment = center_alignment
    ws["B7"].border = thin_border
    ws["C7"].value = "INR(values in Crores/Lakhs)"
    ws["C7"].font = bold_font
    ws["C7"].alignment = center_alignment
    ws["C7"].border = thin_border
    ws["D7"].border = thin_border
    ws["E7"].border = thin_border

    # Income Section
    ws["A8"].value = "Income"
    ws["A8"].font = bold_font
    ws["A8"].alignment = center_alignment
    ws["A8"].border = thin_border

    income_particulars = ["Net Sales", "Change in stock", "Other Income", "Total Income"]
    for i, item in enumerate(income_particulars, start=9):
        ws[f"A{i}"].value = item
        ws[f"A{i}"].border = thin_border
        ws[f"A{i}"].alignment = left_alignment

    # Less Reclassification items
    Less_Reclassification_items_row = 13
    ws["A" + str(Less_Reclassification_items_row)].value = "Less: Reclassification items"
    ws["A" + str(Less_Reclassification_items_row)].font = bold_font
    ws["A" + str(Less_Reclassification_items_row)].alignment = center_alignment
    ws["A" + str(Less_Reclassification_items_row)].border = thin_border
    ws["A" + str(Less_Reclassification_items_row + 1)].value = "Excise duty (forming part of total cost)"
    ws["A" + str(Less_Reclassification_items_row + 1)].border = thin_border
    ws["A" + str(Less_Reclassification_items_row + 1)].alignment = left_alignment
    ws["A" + str(Less_Reclassification_items_row + 2)].value = "Sales tax (forming part of total cost)"
    ws["A" + str(Less_Reclassification_items_row + 2)].border = thin_border
    ws["A" + str(Less_Reclassification_items_row + 2)].alignment = left_alignment

    # Less non-operating items of income
    Less_non_operating_items_of_income_row = 17
    ws["A" + str(Less_non_operating_items_of_income_row)].value = "Less: non-operating items of income"
    ws["A" + str(Less_non_operating_items_of_income_row)].font = bold_font
    ws["A" + str(Less_non_operating_items_of_income_row)].alignment = center_alignment
    ws["A" + str(Less_non_operating_items_of_income_row)].border = thin_border

    non_operating_items = ["Rent received", "Dividend received", "Interest received", 
                           "Write back of diminution in value of investment", "Gain on sale of fixed assets", 
                           "Gain on sale of investments", "Grants/Donations received", "Any other item", "Operating Income"]
    for i, item in enumerate(non_operating_items, start=18):
        ws[f"A{i}"].value = item
        ws[f"A{i}"].border = thin_border
        ws[f"A{i}"].alignment = left_alignment

    # Expenditure section
    expenditure_row = 27
    ws["A" + str(expenditure_row)].value = "II. Expenditure"
    ws["A" + str(expenditure_row)].font = bold_font
    ws["A" + str(expenditure_row)].alignment = center_alignment
    ws["A" + str(expenditure_row)].border = thin_border
    ws["A" + str(expenditure_row + 1)].value = "Total expenditure"
    ws["A" + str(expenditure_row + 1)].border = thin_border
    ws["A" + str(expenditure_row + 1)].alignment = left_alignment

    less_reclassification_items_row = 29
    ws["A" + str(less_reclassification_items_row)].value = "Less: Reclassification items"
    ws["A" + str(less_reclassification_items_row)].font = bold_font
    ws["A" + str(less_reclassification_items_row)].alignment = center_alignment
    ws["A" + str(less_reclassification_items_row)].border = thin_border
    ws["A" + str(less_reclassification_items_row + 1)].value = "Excise duty (forming part of total cost)"
    ws["A" + str(less_reclassification_items_row + 1)].border = thin_border
    ws["A" + str(less_reclassification_items_row + 1)].alignment = left_alignment
    ws["A" + str(less_reclassification_items_row + 2)].value = "Sales tax (forming part of total cost)"
    ws["A" + str(less_reclassification_items_row + 2)].border = thin_border
    ws["A" + str(less_reclassification_items_row + 2)].alignment = left_alignment
    ws["A" + str(less_reclassification_items_row + 3)].value = "Change in stock"
    ws["A" + str(less_reclassification_items_row + 3)].border = thin_border
    ws["A" + str(less_reclassification_items_row + 3)].alignment = left_alignment

    # Less non-operating expenses
    less_non_operating_expenses_row = 33
    ws["A" + str(less_non_operating_expenses_row)].value = "Less: non-operating expenses"
    ws["A" + str(less_non_operating_expenses_row)].font = bold_font
    ws["A" + str(less_non_operating_expenses_row)].alignment = center_alignment
    ws["A" + str(less_non_operating_expenses_row)].border = thin_border
    ws["A" + str(less_non_operating_expenses_row + 1)].value = "Interest paid"
    ws["A" + str(less_non_operating_expenses_row + 1)].border = thin_border
    ws["A" + str(less_non_operating_expenses_row + 1)].alignment = left_alignment
    ws["A" + str(less_non_operating_expenses_row + 2)].value = "Less: Interest expense on lease liability (rentals)"
    ws["A" + str(less_non_operating_expenses_row + 2)].border = thin_border
    ws["A" + str(less_non_operating_expenses_row + 2)].alignment = left_alignment
    ws["A" + str(less_non_operating_expenses_row + 3)].value = "Other Financing charges"
    ws["A" + str(less_non_operating_expenses_row + 3)].border = thin_border
    ws["A" + str(less_non_operating_expenses_row + 3)].alignment = left_alignment
    ws["A" + str(less_non_operating_expenses_row + 4)].value = "Loss on sale of fixed assets"
    ws["A" + str(less_non_operating_expenses_row + 4)].border = thin_border
    ws["A" + str(less_non_operating_expenses_row + 4)].alignment = left_alignment
    ws["A" + str(less_non_operating_expenses_row + 5)].value = "Loss on Sale of Investment"
    ws["A" + str(less_non_operating_expenses_row + 5)].border = thin_border
    ws["A" + str(less_non_operating_expenses_row + 5)].alignment = left_alignment
    ws["A" + str(less_non_operating_expenses_row + 6)].value = "Provision for diminution in the value of investment"
    ws["A" + str(less_non_operating_expenses_row + 6)].border = thin_border
    ws["A" + str(less_non_operating_expenses_row + 6)].alignment = left_alignment
    ws["A" + str(less_non_operating_expenses_row + 7)].value = "Preliminary Expenses Written Off"
    ws["A" + str(less_non_operating_expenses_row + 7)].border = thin_border
    ws["A" + str(less_non_operating_expenses_row + 7)].alignment = left_alignment
    ws["A" + str(less_non_operating_expenses_row + 8)].value = "Donations / CSR"
    ws["A" + str(less_non_operating_expenses_row + 8)].border = thin_border
    ws["A" + str(less_non_operating_expenses_row + 8)].alignment = left_alignment
    ws["A" + str(less_non_operating_expenses_row + 9)].value = "Loss due to fire/theft/natural calamity/breakdown/strike"
    ws["A" + str(less_non_operating_expenses_row + 9)].border = thin_border
    ws["A" + str(less_non_operating_expenses_row + 9)].alignment = left_alignment
    ws["A" + str(less_non_operating_expenses_row + 10)].value = "Voluntary Retirement Scheme"
    ws["A" + str(less_non_operating_expenses_row + 10)].border = thin_border
    ws["A" + str(less_non_operating_expenses_row + 10)].alignment = left_alignment
    ws["A" + str(less_non_operating_expenses_row + 11)].value = "Any other item"
    ws["A" + str(less_non_operating_expenses_row + 11)].border = thin_border
    ws["A" + str(less_non_operating_expenses_row + 11)].alignment = left_alignment
    ws["A" + str(less_non_operating_expenses_row + 12)].value = "Operating Expenses before adjustments"
    ws["A" + str(less_non_operating_expenses_row + 12)].border = thin_border
    ws["A" + str(less_non_operating_expenses_row + 12)].alignment = left_alignment
    ws["A" + str(less_non_operating_expenses_row + 13)].value = "Re-measurement of net defined liability"
    ws["A" + str(less_non_operating_expenses_row + 13)].border = thin_border
    ws["A" + str(less_non_operating_expenses_row + 13)].alignment = left_alignment
    ws["A" + str(less_non_operating_expenses_row + 14)].value = "Operating Expenses after adjustments"
    ws["A" + str(less_non_operating_expenses_row + 14)].border = thin_border
    ws["A" + str(less_non_operating_expenses_row + 14)].alignment = left_alignment

    # Profits section
    profits_row = 48
    ws["A" + str(profits_row)].value = "III. Profits"
    ws["A" + str(profits_row)].font = bold_font
    ws["A" + str(profits_row)].alignment = center_alignment
    ws["A" + str(profits_row)].border = thin_border
    ws["A" + str(profits_row + 1)].value = "Net profits"
    ws["A" + str(profits_row + 1)].border = thin_border
    ws["A" + str(profits_row + 1)].alignment = left_alignment
    ws["A" + str(profits_row + 2)].value = "Operating profits"
    ws["A" + str(profits_row + 2)].border = thin_border
    ws["A" + str(profits_row + 2)].alignment = left_alignment

    # Profit level indicator
    profit_level_indicator_row = 51
    ws["A" + str(profit_level_indicator_row)].value = "IV. Profit Level Indicator"
    ws["A" + str(profit_level_indicator_row)].font = bold_font
    ws["A" + str(profit_level_indicator_row)].alignment = center_alignment
    ws["A" + str(profit_level_indicator_row)].border = thin_border
    ws["A" + str(profit_level_indicator_row + 1)].value = "Operating Profits/Operating Expenses (using foreign exchange fluctuation as operating)"
    ws["A" + str(profit_level_indicator_row + 1)].border = thin_border
    ws["A" + str(profit_level_indicator_row + 1)].alignment = left_alignment

    # Filter computation
    filter_computation_row = 53
    ws["A" + str(filter_computation_row)].value = "Filter computation"
    ws["A" + str(filter_computation_row)].font = bold_font
    ws["A" + str(filter_computation_row)].alignment = center_alignment
    ws["A" + str(filter_computation_row)].border = thin_border

    rpt_filter_row = 54
    ws["A" + str(rpt_filter_row)].value = "1. RPT Filter"
    ws["A" + str(rpt_filter_row)].font = bold_font
    ws["A" + str(rpt_filter_row)].alignment = center_alignment
    ws["A" + str(rpt_filter_row)].border = thin_border
    ws["A" + str(rpt_filter_row + 1)].value = "Nature of transaction"
    ws["A" + str(rpt_filter_row + 1)].border = thin_border
    ws["A" + str(rpt_filter_row + 1)].alignment = left_alignment
    ws["A" + str(rpt_filter_row + 2)].value = "Purchase of goods"
    ws["A" + str(rpt_filter_row + 2)].border = thin_border
    ws["A" + str(rpt_filter_row + 2)].alignment = left_alignment
    ws["A" + str(rpt_filter_row + 3)].value = "Sale of goods"
    ws["A" + str(rpt_filter_row + 3)].border = thin_border
    ws["A" + str(rpt_filter_row + 3)].alignment = left_alignment
    ws["A" + str(rpt_filter_row + 4)].value = "Provision of services"
    ws["A" + str(rpt_filter_row + 4)].border = thin_border
    ws["A" + str(rpt_filter_row + 4)].alignment = left_alignment
    ws["A" + str(rpt_filter_row + 5)].value = "Receipt of services"
    ws["A" + str(rpt_filter_row + 5)].border = thin_border
    ws["A" + str(rpt_filter_row + 5)].alignment = left_alignment
    ws["A" + str(rpt_filter_row + 6)].value = "Payment / receipt of royalty"
    ws["A" + str(rpt_filter_row + 6)].border = thin_border
    ws["A" + str(rpt_filter_row + 6)].alignment = left_alignment
    ws["A" + str(rpt_filter_row + 7)].value = "Other Expenses"
    ws["A" + str(rpt_filter_row + 7)].border = thin_border
    ws["A" + str(rpt_filter_row + 7)].alignment = left_alignment
    ws["A" + str(rpt_filter_row + 8)].value = "Other Incomes"
    ws["A" + str(rpt_filter_row + 8)].border = thin_border
    ws["A" + str(rpt_filter_row + 8)].alignment = left_alignment
    ws["A" + str(rpt_filter_row + 9)].value = "Value of transactions with related parties"
    ws["A" + str(rpt_filter_row + 9)].border = thin_border
    ws["A" + str(rpt_filter_row + 9)].alignment = left_alignment
    ws["A" + str(rpt_filter_row + 10)].value = "% of transactions with related parties"
    ws["A" + str(rpt_filter_row + 10)].border = thin_border
    ws["A" + str(rpt_filter_row + 10)].alignment = left_alignment

    # Employee cost filter
    employee_cost_filter_row = 65
    ws["A" + str(employee_cost_filter_row)].value = "2. Employee cost filter"
    ws["A" + str(employee_cost_filter_row)].font = bold_font
    ws["A" + str(employee_cost_filter_row)].alignment = center_alignment
    ws["A" + str(employee_cost_filter_row)].border = thin_border
    ws["A" + str(employee_cost_filter_row + 1)].value = "Employee benefit expense"
    ws["A" + str(employee_cost_filter_row + 1)].border = thin_border
    ws["A" + str(employee_cost_filter_row + 1)].alignment = left_alignment
    ws["A" + str(employee_cost_filter_row + 2)].value = "Employee benefit expense as a % of total expenses"
    ws["A" + str(employee_cost_filter_row + 2)].border = thin_border
    ws["A" + str(employee_cost_filter_row + 2)].alignment = left_alignment

    # Export revenue filter
    export_revenue_filter_row = 68
    ws["A" + str(export_revenue_filter_row)].value = "3. Export revenue filter"
    ws["A" + str(export_revenue_filter_row)].font = bold_font
    ws["A" + str(export_revenue_filter_row)].alignment = center_alignment
    ws["A" + str(export_revenue_filter_row)].border = thin_border
    ws["A" + str(export_revenue_filter_row + 1)].value = "Export revenue / Revenue earned in Foreign exchange"
    ws["A" + str(export_revenue_filter_row + 1)].border = thin_border
    ws["A" + str(export_revenue_filter_row + 1)].alignment = left_alignment
    ws["A" + str(export_revenue_filter_row + 2)].value = "Export revenue as a % of total revenue"
    ws["A" + str(export_revenue_filter_row + 2)].border = thin_border
    ws["A" + str(export_revenue_filter_row + 2)].alignment = left_alignment

    # Information for working capital adjustment
    ws["A72"].value = "B. Information for working capital adjustment"
    ws["A72"].font = bold_font
    ws["A72"].alignment = center_alignment
    ws["A72"].border = thin_border
    ws["A73"].value = "Particulars"
    ws["A73"].font = bold_font
    ws["A73"].alignment = center_alignment
    ws["A73"].border = thin_border
    ws["B73"].value = "FY 2022-23"
    ws["B73"].font = bold_font
    ws["B73"].alignment = center_alignment
    ws["B73"].border = thin_border
    ws["C73"].value = "FY 2021-22"
    ws["C73"].font = bold_font
    ws["C73"].alignment = center_alignment
    ws["C73"].border = thin_border
    ws["D73"].value = "FY 2020-21"
    ws["D73"].font = bold_font
    ws["D73"].alignment = center_alignment
    ws["D73"].border = thin_border
    ws["E73"].value = "FY 2019-20"
    ws["E73"].font = bold_font
    ws["E73"].alignment = center_alignment
    ws["E73"].border = thin_border
    ws["B74"].value = "INR"
    ws["B74"].font = bold_font
    ws["B74"].alignment = center_alignment
    ws["B74"].border = thin_border
    ws["C74"].value = "INR"
    ws["C74"].font = bold_font
    ws["C74"].alignment = center_alignment
    ws["C74"].border = thin_border
    ws["D74"].value = "INR"
    ws["D74"].font = bold_font
    ws["D74"].alignment = center_alignment
    ws["D74"].border = thin_border
    ws["E74"].value = "INR"
    ws["E74"].font = bold_font
    ws["E74"].alignment = center_alignment
    ws["E74"].border = thin_border

    section_b_row = 75
    ws["A" + str(section_b_row)].value = "Sundry Debtors/Bills receivable"
    ws["A" + str(section_b_row)].border = thin_border
    ws["A" + str(section_b_row)].alignment = left_alignment
    ws["A" + str(section_b_row + 1)].value = "Unbilled Revenue"
    ws["A" + str(section_b_row + 1)].border = thin_border
    ws["A" + str(section_b_row + 1)].alignment = left_alignment
    ws["A" + str(section_b_row + 2)].value = "Less: Advances from customers/Unearned revenue"
    ws["A" + str(section_b_row + 2)].border = thin_border
    ws["A" + str(section_b_row + 2)].alignment = left_alignment
    ws["A" + str(section_b_row + 3)].value = "Less: Advance Revenue/Prepaid revenue"
    ws["A" + str(section_b_row + 3)].border = thin_border
    ws["A" + str(section_b_row + 3)].alignment = left_alignment
    ws["A" + str(section_b_row + 4)].value = "Accounts receivable"
    ws["A" + str(section_b_row + 4)].border = thin_border
    ws["A" + str(section_b_row + 4)].alignment = left_alignment
    ws["A" + str(section_b_row + 5)].value = "Sundry Creditors/Bills Payable"
    ws["A" + str(section_b_row + 5)].border = thin_border
    ws["A" + str(section_b_row + 5)].alignment = left_alignment
    ws["A" + str(section_b_row + 6)].value = "Less: Prepaid Expenses"
    ws["A" + str(section_b_row + 6)].border = thin_border
    ws["A" + str(section_b_row + 6)].alignment = left_alignment
    ws["A" + str(section_b_row + 7)].value = "Less: Advances paid to vendors"
    ws["A" + str(section_b_row + 7)].border = thin_border
    ws["A" + str(section_b_row + 7)].alignment = left_alignment
    ws["A" + str(section_b_row + 8)].value = "Less: Advance recoverable in cash or in kind"
    ws["A" + str(section_b_row + 8)].border = thin_border
    ws["A" + str(section_b_row + 8)].alignment = left_alignment
    ws["A" + str(section_b_row + 9)].value = "Accounts payable"
    ws["A" + str(section_b_row + 9)].border = thin_border
    ws["A" + str(section_b_row + 9)].alignment = left_alignment
    ws["A" + str(section_b_row + 10)].value = "Raw Material and Components"
    ws["A" + str(section_b_row + 10)].border = thin_border
    ws["A" + str(section_b_row + 10)].alignment = left_alignment
    ws["A" + str(section_b_row + 11)].value = "Work in Progress"
    ws["A" + str(section_b_row + 11)].border = thin_border
    ws["A" + str(section_b_row + 11)].alignment = left_alignment
    ws["A" + str(section_b_row + 12)].value = "Finished Goods"
    ws["A" + str(section_b_row + 12)].border = thin_border
    ws["A" + str(section_b_row + 12)].alignment = left_alignment
    ws["A" + str(section_b_row + 13)].value = "Inventory"
    ws["A" + str(section_b_row + 13)].border = thin_border
    ws["A" + str(section_b_row + 13)].alignment = left_alignment
    ws["A" + str(section_b_row + 14)].value = "Net working capital"
    ws["A" + str(section_b_row + 14)].border = thin_border
    ws["A" + str(section_b_row + 14)].alignment = left_alignment

    # Additional cell styling as per user request
    special_cells = ["B26", "C26", "B45", "C45", "B47", "C47", "B49", "C49", 
                     "B50", "C50", "B52", "C52", "B63", "C63", "B79", "C79", 
                     "B84", "C84", "B88", "C88", "B89", "C89"]

    for cell in special_cells:
        ws[cell].border = thick_border
        ws[cell].fill = light_grey_fill
        ws[cell].alignment = center_alignment

    # Save the workbook
    folder_name = os.path.basename(folder)
    excel_path = os.path.join(excel_folder, f"{folder_name}.xlsx")
    new_wb.save(excel_path)
    return excel_path






def fill_values(check_excel, fill_excel):
    # Load the workbooks
    checking = load_workbook(filename=check_excel)
    filling = load_workbook(filename=fill_excel)

    # Assume we're working with the first sheet in both workbooks
    check_sheet = checking.active
    fill_sheet = filling.active

    # Column headers to check in the check_sheet
    checking_columns = [
        'Total revenue from operations',
        'Total other income',
        'Total dividend income',
        'Total interest income',
        'Total employee benefit expense',
        'Total other non-operating income',
        'Total expenses',
        'Total finance costs',
        'Surplus on disposal, discard, demolishment and destruction of depreciable property, plant and equipment',
        'Total net gain/loss on sale of investments',
        'Interest lease financing',
        'Other interest charges',
        'Aggregate provision for diminution in value of current investments',
        'CSR expenditure',
        'Purchase of Goods / Services',
        'Total trade receivables',
        'Trade payables, current',
        'Other current assets others',
        'Loss on disposal, discard, demolishment and destruction of depreciable property plant and equipment',
        'Total net loss on sale of investments',
        'Other current financial liabilities',
        'Other current financial liabilities',
        'Other current financial assets others',
        'Other current financial assets others',
        'Other current financial assets others'
    ]

    # Dictionary to store values from check_sheet
    extracted_values = {}

    # Function to get value from check_sheet based on the header
    def get_value():
        for row in check_sheet.iter_rows(min_row=2, values_only=True):  # Iterate over each row starting from the second
            header = row[0]
            if header in checking_columns:
                # Check if the header already exists in extracted_values
                if header not in extracted_values:
                    extracted_values[header] = (row[1:])
                else:
                    # Find a new unique header name
                    index = 1
                    new_header = f"{header}{index}"
                    while new_header in extracted_values:
                        index += 1
                        new_header = f"{header}{index}"
                    extracted_values[new_header] = (row[1:])
                    
                print(f"Extracted values for '{header}': {extracted_values[header]}")
    get_value()

    B9, C9, B10, C10, B11, C11 = 0, 0, 0, 0, 0, 0

    # Function to set value in fill_sheet
    def set_value(cell, value):
        fill_sheet[cell].value = value
        fill_sheet[cell].alignment = Alignment(horizontal='right')  # Align text to the right
        print(f"Set value '{value}' in cell '{cell}'")  # Debugging statement

    # Clean up the extracted values
    def clean_value(value):
        return float(re.sub(r'[^\d.]', '', str(value)))

    # Map the extracted values to specific cells in fill_sheet
    if 'Total revenue from operations' in extracted_values:
        total_revenue_from_operations = extracted_values['Total revenue from operations']
        B9 = clean_value(total_revenue_from_operations[0])
        C9 = clean_value(total_revenue_from_operations[1])
        set_value("B9", B9)
        set_value("C9", C9)

    if 'Total other income' in extracted_values:
        total_other_income = extracted_values['Total other income']
        B11 = clean_value(total_other_income[0])
        C11 = clean_value(total_other_income[1])
        set_value("B11", B11)
        set_value("C11", C11)

    if 'Total dividend income' in extracted_values:
        total_dividend_income = extracted_values['Total dividend income']
        set_value("B19", clean_value(total_dividend_income[0]))
        set_value("C19", clean_value(total_dividend_income[1]))

    if 'Total interest income' in extracted_values:
        total_interest_income = extracted_values['Total interest income']
        set_value("B20", clean_value(total_interest_income[0]))
        set_value("C20", clean_value(total_interest_income[1]))

    if 'Total finance costs' in extracted_values:
        total_finance_costs = extracted_values['Total finance costs']
        set_value("B34", clean_value(total_finance_costs[0]))
        set_value("C34", clean_value(total_finance_costs[1]))

    if 'Total employee benefit expense' in extracted_values:
        total_employee_benefit = extracted_values['Total employee benefit expense']
        set_value("B66", clean_value(total_employee_benefit[0]))
        set_value("C66", clean_value(total_employee_benefit[1]))

    if 'Surplus on disposal, discard, demolishment and destruction of depreciable property, plant and equipment' in extracted_values:
        surplus = extracted_values['Surplus on disposal, discard, demolishment and destruction of depreciable property, plant and equipment']
        set_value("B22", clean_value(surplus[0]))
        set_value("C22", clean_value(surplus[1]))

    if 'Total net gain/loss on sale of investments' in extracted_values:
        total_net_gain = extracted_values['Total net gain/loss on sale of investments']
        set_value("B23", clean_value(total_net_gain[0]))
        set_value("C23", clean_value(total_net_gain[1]))

    if 'Total expenses' in extracted_values:
        expenses = extracted_values['Total expenses']
        set_value("B28", clean_value(expenses[0]))
        set_value("C28", clean_value(expenses[1]))

    if 'Interest lease financing' in extracted_values:
        lease = extracted_values['Interest lease financing']
        set_value("B35", clean_value(lease[0]))
        set_value("C35", clean_value(lease[1]))

    if 'Other interest charges' in extracted_values:
        interest_charges = extracted_values['Other interest charges']
        set_value("B36", clean_value(interest_charges[0]))
        set_value("C36", clean_value(interest_charges[1]))

    if 'Loss on disposal, discard, demolishment and destruction of depreciable property plant and equipment' in extracted_values:
        loss_on = extracted_values['Loss on disposal, discard, demolishment and destruction of depreciable property plant and equipment']
        set_value("B37", clean_value(loss_on[0]))
        set_value("C37", clean_value(loss_on[1]))

    if 'Total net loss on sale of investments' in extracted_values:
        net_loss = extracted_values['Total net loss on sale of investments']
        set_value("B38", clean_value(net_loss[0]))
        set_value("C38", clean_value(net_loss[1]))

    if 'Aggregate provision for diminution in value of current investments' in extracted_values:
        aggregate = extracted_values['Aggregate provision for diminution in value of current investments']
        set_value("B39", clean_value(aggregate[0]))
        set_value("C39", clean_value(aggregate[1]))

    if 'CSR expenditure' in extracted_values:
        csr = extracted_values['CSR expenditure']
        set_value("B41", clean_value(csr[0]))
        set_value("C41", clean_value(csr[1]))

    if 'Purchase of Goods / Services' in extracted_values:
        purchase = extracted_values['Purchase of Goods / Services']
        set_value("B56", clean_value(purchase[0]))
        set_value("C56", clean_value(purchase[1]))

    if 'Total trade receivables' in extracted_values:
        trade = extracted_values['Total trade receivables']
        set_value("B75", clean_value(trade[0]))
        set_value("C75", clean_value(trade[1]))

    if 'Trade payables, current' in extracted_values:
        trade = extracted_values['Trade payables, current']
        set_value("B80", clean_value(trade[0]))
        set_value("C80", clean_value(trade[1]))

    if 'Other current financial assets others' in extracted_values:
        current = extracted_values['Other current financial assets others']
        set_value("B76", clean_value(current[0]))
        set_value("C76", clean_value(current[1]))

    if 'Other current financial liabilities' in extracted_values:
        other = extracted_values['Other current financial liabilities']
        set_value("B77", clean_value(other[0]))
        set_value("C77", clean_value(other[1]))

    if 'Other current financial assets others1' in extracted_values:
        fincaiacl = extracted_values['Other current financial assets others1']
        set_value("B76", clean_value(fincaiacl[0]))
        set_value("C76", clean_value(fincaiacl[1]))

    if 'Other current financial assets others2' in extracted_values:
        tree = extracted_values['Other current financial assets others2']
        set_value("B82", clean_value(tree[0]))
        set_value("C82", clean_value(tree[1]))

    if 'Other current financial assets others3' in extracted_values:
        tree = extracted_values['Other current financial assets others3']
        set_value("B83", clean_value(tree[0]))
        set_value("C83", clean_value(tree[1]))

    if 'Total other non-operating income' in extracted_values:
        value = extracted_values["Total other non-operating income"]
        set_value("B25", clean_value(value[0]))
        set_value("C25", clean_value(value[1]))

    # Set summed values
    if (fill_sheet["B12"].value is None and fill_sheet["C12"].value is None) or (fill_sheet["B12"].value == 0 and fill_sheet["C12"].value == 0):
        set_value("B12", B9 + B10 + B11)
        set_value("C12", C9 + C10 + C11)

    # Save the filled workbook
    filling.save(filename=fill_excel)


def update_values(fill_excel, result, folder_path):
    filling = load_workbook(filename=fill_excel)
    folder_name = os.path.basename(folder_path)

    for i in range(len(result)):
        new_name = folder_name + ".pdf"
        print(new_name, result[i][0])
        if new_name == result[i][0]:
            arr = result.pop(i)
            break

    fill_sheet = filling.active

    def get_value(cell):
        return fill_sheet[cell].value

    def set_value(cell, value):
        fill_sheet[cell].value = value
        fill_sheet[cell].alignment = Alignment(horizontal='right')  # Align text to the right
        print(f"Set value '{value}' in cell '{cell}'")

    set_value("B9", get_value("B9") * arr[2])
    set_value("C9", get_value("C9") * arr[2])
    set_value("B11", get_value("B11") * arr[2])
    set_value("C11", get_value("C11") * arr[2])
    set_value("B19", get_value("C19") * arr[2])
    set_value("C19", get_value("C19") * arr[2])
    set_value("B20", get_value("B20") * arr[2])
    set_value("C20", get_value("C20") * arr[2])
    set_value("B34", get_value("B34") * arr[2])
    set_value("C34", get_value("C34") * arr[2])
    set_value("B66", get_value("B66") * arr[2])
    set_value("C66", get_value("C66") * arr[2])

    set_value("B12", get_value("B12") * arr[2])
    set_value("C12", get_value("C12") * arr[2])

    set_value("B1", arr[1])

    other_cells = [
        "B10", "C10", "B22", "C22", "B23", "C23", "B28", "C28", 
        "B35", "C35", "B36", "C36", "B37", "C37", "B38", "C38", 
        "B39", "C39", "B41", "C41", "B56", "C56", "B75", "C75", 
        "B80", "C80", "B76", "C76", "B77", "C77", "B82", "C82", 
        "B83", "C83", "B25", "C25"
    ]
    
    for cell in other_cells:
        print(cell, get_value(cell))
        original_value = get_value(cell)
        if original_value is not None:
            set_value(cell, original_value * arr[2])

    def get_value_with_default(cell, default=0):
        value = get_value(cell)
        return value if value is not None else default

    set_value("B26", get_value_with_default("B12") - (get_value_with_default("B14") + get_value_with_default("B15") + get_value_with_default("B16") + get_value_with_default("B18") + get_value_with_default("B19") + get_value_with_default("B20") + get_value_with_default("B21") + get_value_with_default("B22") + get_value_with_default("B23") + get_value_with_default("B24") + get_value_with_default("B25")))
    set_value("C26", get_value_with_default("C12") - (get_value_with_default("C14") + get_value_with_default("C15") + get_value_with_default("C16") + get_value_with_default("C18") + get_value_with_default("C19") + get_value_with_default("C20") + get_value_with_default("C21") + get_value_with_default("C22") + get_value_with_default("C23") + get_value_with_default("C24") + get_value_with_default("C25")))

    set_value("B45", (get_value_with_default("B28") + get_value_with_default("B35")) - (get_value_with_default("B30") + get_value_with_default("B31") + get_value_with_default("B32") + get_value_with_default("B34") + get_value_with_default("B36") + get_value_with_default("B37") + get_value_with_default("B38") + get_value_with_default("B39") + get_value_with_default("B40") + get_value_with_default("B41") + get_value_with_default("B42") + get_value_with_default("B43") + get_value_with_default("B44")))
    set_value("C45", (get_value_with_default("C28") + get_value_with_default("C35")) - (get_value_with_default("C30") + get_value_with_default("C31") + get_value_with_default("C32") + get_value_with_default("C34") + get_value_with_default("C36") + get_value_with_default("C37") + get_value_with_default("C38") + get_value_with_default("C39") + get_value_with_default("C40") + get_value_with_default("C41") + get_value_with_default("C42") + get_value_with_default("C43") + get_value_with_default("C44")))

    set_value("B47", get_value_with_default("B45") - get_value_with_default("B46"))
    set_value("C47", get_value_with_default("C45") - get_value_with_default("C46"))

    set_value("B49", get_value_with_default("B12") - get_value_with_default("B28"))
    set_value("C49", get_value_with_default("C12") - get_value_with_default("C28"))

    set_value("B50", get_value_with_default("B26") - get_value_with_default("B47"))
    set_value("C50", get_value_with_default("C26") - get_value_with_default("C47"))

    set_value("B52", get_value_with_default("B50") / get_value_with_default("B47") * 100)
    set_value("C52", get_value_with_default("C50") / get_value_with_default("C47") * 100)

    set_value("B63", get_value_with_default("B52") + get_value_with_default("B62"))
    set_value("C63", get_value_with_default("C52") + get_value_with_default("C62"))

    set_value("B79", get_value_with_default("B75") + get_value_with_default("B76") - get_value_with_default("B77") - get_value_with_default("B78"))
    set_value("C79", get_value_with_default("C75") + get_value_with_default("C76") - get_value_with_default("C77") - get_value_with_default("C78"))

    set_value("B84", get_value_with_default("B80") - get_value_with_default("B81") - get_value_with_default("B82") - get_value_with_default("B83"))
    set_value("C84", get_value_with_default("C80") - get_value_with_default("C81") - get_value_with_default("C82") - get_value_with_default("C83"))

    set_value("B88", get_value_with_default("B85") + get_value_with_default("B87") + get_value_with_default("B86"))
    set_value("C88", get_value_with_default("C85") + get_value_with_default("C87") + get_value_with_default("C86"))

    set_value("B89", get_value_with_default("B79") + get_value_with_default("B88") - get_value_with_default("B84"))
    set_value("C89", get_value_with_default("C79") + get_value_with_default("C88") - get_value_with_default("C84"))

    filling.save(filename=fill_excel)

'''
result = [['Name.pdf', 'ECLERX', 1000000]]
fill_values('/home/geetansh/Desktop/Complete_Project/FinDetector/output_path/XBRL_financial_statements_duly_authenticated_as_per_section_134_including_Board/500100NotesSubclassificationand.xlsx', 
            'financial_template_stylish.xlsx', result)
'''
def combine_excel_files(input_dir, output_file):
    try:
        output_path = os.path.join(input_dir, output_file)
        
        with xlsxwriter.Workbook(output_path, {'nan_inf_to_errors': True}) as final_workbook:
            header_format = final_workbook.add_format({'bold': True, 'border': 1})
            data_format = final_workbook.add_format({'border': 1})
            title_format = final_workbook.add_format({'bold': True, 'border': 1, 'align': 'center'})

            for subfolder in os.listdir(input_dir):
                subfolder_path = os.path.join(input_dir, subfolder)
                
                if os.path.isdir(subfolder_path):
                    row_idx = 0
                    
                    # Create a new sheet for each subfolder
                    sheet_name = re.sub(r'[\W_]+', '', subfolder)[:31]  # Clean sheet name and truncate to 31 chars
                    final_worksheet = final_workbook.add_worksheet(sheet_name)
                    final_worksheet.set_column('A:Z', 50)  # Set column width
                    
                    for file_name in os.listdir(subfolder_path):
                        if file_name.endswith('.xlsx'):
                            file_path = os.path.join(subfolder_path, file_name)
                            df = pd.read_excel(file_path)
                            
                            # Remove numbers from the beginning of the file name and extension
                            cleaned_file_name = re.sub(r'^\d+', '', file_name).strip()
                            cleaned_file_name = re.sub(r'\.xlsx$', '', cleaned_file_name)
                            
                            # Write the cleaned name of the Excel file as a title for the table
                            final_worksheet.merge_range(row_idx, 0, row_idx, len(df.columns) - 1, cleaned_file_name, title_format)
                            row_idx += 1
                            
                            # Write the header of the table with bold format and borders
                            for col_num, value in enumerate(df.columns.values):
                                final_worksheet.write(row_idx, col_num, value, header_format)
                            row_idx += 1
                            
                            # Write the data rows with borders
                            for index, row in df.iterrows():
                                for col_num, value in enumerate(row):
                                    # Replace null values or null characters with an empty string
                                    if pd.isna(value) or value == '\0':
                                        value = ''
                                    final_worksheet.write(row_idx, col_num, value, data_format)
                                row_idx += 1
                            
                            # Group the rows and name the combined table "margin template"
                            final_worksheet.set_row(row_idx - len(df) - 1, None, None, {'collapsed': True})
                            final_worksheet.set_row(row_idx - len(df), None, None, {'level': 1, 'hidden': True})
                            for r in range(row_idx - len(df) + 1, row_idx):
                                final_worksheet.set_row(r, None, None, {'level': 1, 'hidden': True})
                            
                            # Set a named range for the combined table
                            final_workbook.define_name('margin_template', f'{sheet_name}!$A${row_idx - len(df) - 1}:$A${row_idx}')
                            
                            row_idx += 1
                    
                    # Add an empty group named "calculated data" at the end
                    final_worksheet.merge_range(row_idx, 0, row_idx, len(df.columns) - 1, 'Calculated Data', title_format)
                    final_worksheet.set_row(row_idx, None, None, {'collapsed': True})
                    row_idx += 1

                    final_worksheet.outline_settings(True, False, False, False)
        
        print(f"Combined data saved to {output_file}")
    except Exception as e:
        print(f"Error in combine_excel_files: {e}")