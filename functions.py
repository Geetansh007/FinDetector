import os
import shutil
from werkzeug.utils import secure_filename
from extract import PDFExtractor, Save
from extract_excel import fill_values, create_excel_template, update_values
import glob
import zipfile
import pandas as pd
import re
from openpyxl.styles import Font,Border,Alignment,Side,PatternFill
from openpyxl import load_workbook,Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() == 'pdf'

def clear_directories(uploads, output_base_folder, excel_folder):
    """
    Remove specified directories to ensure a fresh start for each upload process.
    """
    for folder in [uploads, output_base_folder, excel_folder]:
        if os.path.exists(folder):
            shutil.rmtree(folder)
            print(f"Cleared existing directory: {folder}")

def save_uploaded_files(request, upload_folder='uploads', output_base_folder='output', excel_folder='Excel_folder'):
    try:
        if request.method != 'POST':
            raise ValueError("Invalid request method")

        os.makedirs(upload_folder, exist_ok=True)

        if 'files[]' not in request.files:
            raise ValueError("No files part in the request")

        files = request.files.getlist('files[]')

        if not files:
            raise ValueError("No files selected for uploading")

        for file in files:
            if file.filename == '':
                raise ValueError("No selected file")
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                file.save(os.path.join(upload_folder, filename))
            else:
                raise ValueError("Invalid file format. Only PDF files are allowed.")

        return upload_folder

    except Exception as e:
        raise e

def process_uploaded_pdfs(upload_folder, output_base_folder):
    os.makedirs(output_base_folder, exist_ok=True)
    results = []
    for filename in os.listdir(upload_folder):
        if filename.endswith('.pdf'):
            file_path = os.path.join(upload_folder, filename)
            output_dir = os.path.join(output_base_folder, os.path.splitext(filename)[0])
            if not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)
                extractor = PDFExtractor(file_path, output_dir)
                extractor.extract_all_tables()

    for filename in os.listdir(upload_folder):
        if filename.endswith('.pdf'):
            file_path = os.path.join(upload_folder, filename)
            company_name, monetary_unit_value = extract_and_save(file_path)
            results.append((filename, company_name, monetary_unit_value))
            print("\n", results)


    return output_base_folder, results

def load_pdf_excel(output_base_path, excel_folder, result):
    try:
        directories = [d for d in os.listdir(output_base_path) if os.path.isdir(os.path.join(output_base_path, d))]

        os.makedirs(excel_folder, exist_ok=True)

        for folder in directories:
            folder_path = os.path.join(output_base_path, folder)
            print(f"Loading files from folder: {folder_path}")

            files = os.listdir(folder_path)
            excel_path = create_excel_template(folder_path, excel_folder)
            print("\nMaking excel\n")
            for file in files:
                if file.lower().endswith('.xlsx'):
                    file_path = os.path.join(folder_path, file)
                    fill_values(file_path, excel_path)
            update_values(excel_path, result,folder_path)

    except Exception as e:
        print(f"An error occurred: {e}")

def extract_and_save(file_path):
    extractor = Save(file_path)
    company_name = extractor.extract_company_name()
    monetary_unit_value = extractor.extract_monetary_unit()
    return company_name, monetary_unit_value

def download_folder():
    folder_name = "Excel_folder"
    zip_file_name = "excel_output.zip"

    with zipfile.ZipFile(zip_file_name, 'w') as zipf:
        excel_files = glob.glob(os.path.join(folder_name, "*.xlsx"))

        for file in excel_files:
            zipf.write(file, os.path.basename(file))

    return zip_file_name

def table_display(folder_path):
    if not os.path.exists(folder_path):
        raise ValueError(f"The folder {folder_path} does not exist.")
    dataframes = {}
    for filename in os.listdir(folder_path):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(folder_path, filename)
            df = pd.read_excel(file_path)
            name = os.path.splitext(filename)[0]
            dataframes[name] = df
            
    return dataframes

def append_data_to_excel(original_excel_path, excel_path):
    try:
        # Load existing workbook and sheets
        existing_sheets = pd.read_excel(original_excel_path, sheet_name=None)
        workbook = load_workbook(original_excel_path)

        # Define styles
        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        thick_border = Border(left=Side(style='thick'), 
                              right=Side(style='thick'), 
                              top=Side(style='thick'), 
                              bottom=Side(style='thick'))

        # Process each sheet and append new data
        for sheet_name, df in existing_sheets.items():
            print(f"Processing sheet: {sheet_name}")
            for filename in os.listdir(excel_path):
                if filename.endswith(".xlsx") or filename.endswith(".xls"):
                    file_name_without_extension = os.path.splitext(filename)[0]
                    file_name_without_underscores = file_name_without_extension.replace('_', '')
                    shortened_name = file_name_without_underscores[:31]
                    
                    if shortened_name == sheet_name:
                        print(shortened_name, sheet_name)
                        shortened_df = pd.read_excel(os.path.join(excel_path, filename))
                        worksheet = workbook[sheet_name]
                        startrow = worksheet.max_row + 6  # Calculate the starting row for new data

                        # Insert the new group label
                        margien_sheet = f"{shortened_name}_new_group"
                        cell = worksheet.cell(row=startrow, column=1, value=margien_sheet)
                        cell.font = bold_font
                        cell.alignment = center_alignment
                        cell.border = thick_border
                        worksheet.column_dimensions['A'].width = 50
                        worksheet.column_dimensions['B'].width = 50
                        worksheet.column_dimensions['C'].width = 50
                        worksheet.column_dimensions['D'].width = 50
                        worksheet.column_dimensions['E'].width = 50

                        # Write new data
                        for r_idx, row in enumerate(dataframe_to_rows(shortened_df, index=False, header=True), start=startrow + 1):
                            for c_idx, value in enumerate(row, 1):
                                cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
                                cell.border = thin_border
                                cell.alignment = center_alignment if r_idx == startrow + 1 else left_alignment

                        # Apply grouping and set named range
                        title_row = startrow
                        header_row = startrow + 1
                        data_start_row = startrow + 2
                        data_end_row = startrow + 1 + len(shortened_df)
                        worksheet.row_dimensions[title_row].outlineLevel = 1
                        worksheet.row_dimensions[header_row].outlineLevel = 1
                        for r in range(data_start_row, data_end_row + 1):
                            worksheet.row_dimensions[r].outlineLevel = 1

                        workbook.create_named_range(f'{sheet_name}_margin_template', worksheet, f'{sheet_name}!$A${title_row + 1}:$A${data_end_row + 1}')

                        break

        # Save the workbook
        workbook.save(original_excel_path)
        print(f"Data appended and workbook saved to {original_excel_path}")

    except Exception as e:
        print(f"Error in append_data_to_excel: {e}")
